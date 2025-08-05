from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAdminUser
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework import status

from Payments.models import Payment
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.utils import timezone

import io
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm


from app .permissions import HasModelPermission 

class AdminPaymentExcelExportAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAdminUser , HasModelPermission]

    def get(self, request):
        status_filter = request.query_params.get("status")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        payments = Payment.objects.select_related('user', 'payment_gateway').all()

        if status_filter:
            payments = payments.filter(status=status_filter)
        if start_date:
            payments = payments.filter(payment_date__date__gte=parse_date(start_date))
        if end_date:
            payments = payments.filter(payment_date__date__lte=parse_date(end_date))

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Payments Report"

        headers = ["ID", "User", "Amount", "Currency", "Status", "Payment Date", "Gateway"]
        sheet.append(headers)

        for payment in payments:
            row = [
                payment.id,
                payment.user.email,
                str(payment.amount),
                payment.currency,
                payment.status,
                payment.payment_date.strftime('%Y-%m-%d %H:%M'),
                payment.payment_gateway.name if payment.payment_gateway else "-"
            ]
            sheet.append(row)

        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 20

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="payments_report_{timezone.now().date()}.xlsx"'
        return response
